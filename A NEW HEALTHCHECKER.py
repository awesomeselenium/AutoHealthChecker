from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import time

global driver

delay=2

def START():
    Options().add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36")
    global driver 
    driver = webdriver.Chrome('chromedriver')
    driver.get("https://hcs.eduro.go.kr/#/loginHome")
    time.sleep(1)
    
def AHC(name,birth,city,level,schoolname,password):
    time.sleep(delay)
    driver.find_element_by_xpath('//*[@id="btnConfirm2"]').click() #다음 페이지로 넘어감
    driver.implicitly_wait(10) 
    time.sleep(delay)
    driver.find_element_by_xpath('//*[@id="WriteInfoForm"]/table/tbody/tr[2]/td/input').send_keys(name) #이름 입력
    driver.find_element_by_xpath('//*[@id="WriteInfoForm"]/table/tbody/tr[3]/td/input').send_keys(birth) #생일 입력
    driver.find_element_by_xpath('//*[@id="WriteInfoForm"]/table/tbody/tr[1]/td/button').click() #학교찾는 버튼 클릭 
    Select(driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[1]/table/tbody/tr[1]/td/select')).select_by_value(city) #시/도 선택
    Select(driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[1]/table/tbody/tr[2]/td/select')).select_by_value(level) #학교급 선택
    driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[1]/table/tbody/tr[3]/td[1]/input').send_keys(schoolname) #학교 이름 작성
    driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[1]/table/tbody/tr[3]/td[2]/button').click() #학교 검색 버튼 클릭
    time.sleep(delay)
    driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[1]/ul/li[1]/span').click() #학교 선택
    driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[2]/input').click()
    time.sleep(delay)
    driver.find_element_by_xpath('//*[@id="btnConfirm"]').click()
    driver.implicitly_wait(10)
    time.sleep(delay)
    driver.find_element_by_xpath('//*[@id="WriteInfoForm"]/table/tbody/tr/td/input').send_keys(password)#비밀번호 입력 
    driver.find_element_by_xpath('//*[@id="btnConfirm"]').click()
    driver.implicitly_wait(10)
    time.sleep(delay)
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/section[2]/div[2]/ul/li/a/span[1]').click() #상위 사용자 선택    
    driver.implicitly_wait(10)
    time.sleep(delay)
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/div/div[2]/div[2]/dl[1]/dd/ul/li[1]/label').click() #1~5번 항목 확인
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/div/div[2]/div[2]/dl[2]/dd/ul/li[1]/label').click()
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/div/div[2]/div[2]/dl[3]/dd/ul/li[1]/label').click()
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/div/div[2]/div[2]/dl[4]/dd/ul/li[1]/label').click()
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/div/div[2]/div[2]/dl[5]/dd/ul/li[1]/label').click()
    driver.find_element_by_xpath('//*[@id="btnConfirm"]').click() #제출 버튼 클릭
    time.sleep(delay)
    driver.implicitly_wait(10)
    print(load_ws.cell(cnt,1).value)
    print(cnt,"/",lastnumber-1,"completed")
    load_ws.cell(row=cnt,column=3).value = 1
    load_wb.save(r'studentlist.xlsx')


def RER():  #정보 클리어
    driver.refresh()
    time.sleep(2)
    driver.switch_to.alert.accept()
    time.sleep(1)
    driver.find_element_by_xpath('/html/body/app-root/div/div[1]/div/button').click()    
    time.sleep(1)
    driver.switch_to.alert.accept()
    time.sleep(1)

def Cycle(name,birth,city,level,schoolname,password): #이름, 생일, 시/도, 학교급, 학교이름, 비밀번호 받아서 AHC와 RER을 수행
    AHC(name,birth,city,level,schoolname,password)
    RER()

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
#프로그램 구동부
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
startnumber_manual=1
load_wb = load_workbook(r'studentlist.xlsx',data_only=True)
load_ws = load_wb['Sheet1']
lastnumber = 1
autostart = 1

START()

#Dictionary 시/도, 학교급 데이터 대응
dic_city = {'서울특별시':'01','부산광역시':'02','대구광역시':'03','인천광역시':'04','광주광역시':'05','대전광역시':'06','울산광역시':'07','세종특별자치시':'08','경기도':'10','강원도':'11','충청북도':'12','충청남도':'13','전라북도':'14','전라남도':'15','경상북도':'16','경상남도':'17','제주특별자치도':'18'}
dic_level = {'유치원':'1','초등학교':'2','중학교':'3','고등학교':'4','특수학교':'5'}
        

while (load_ws.cell(lastnumber,1).value != None):
    lastnumber=lastnumber+1
while (load_ws.cell(autostart,3).value != None):
    autostart=autostart+1
print("StartPoint:", autostart)
startnumber_manual= autostart
print("엑셀 파일의 학생은 자신의 건강상태와 주위 사람들의 건강 상태를 점검하여서 문제가 없었음을 인정하였습니다. ")

for cnt in range(startnumber_manual,lastnumber):
    if (load_ws.cell(cnt,4).value == None): #시/도 있으면 그대로 쓰고 없으면 1행 값
        city = dic_city[load_ws.cell(1,4).value]
    if (load_ws.cell(cnt,4).value != None):
        city = dic_city[load_ws.cell(cnt,4).value]

    if (load_ws.cell(cnt,5).value == None): #학교급 있으면 그대로 쓰고 없으면 1행 값
        level = dic_level[load_ws.cell(1,5).value]
    if (load_ws.cell(cnt,5).value != None):
        level = dic_level[load_ws.cell(cnt,5).value]
        
    if (load_ws.cell(cnt,6).value == None): #학교 이름 있으면 그대로 쓰고 없으면 1행 값
        schoolname = load_ws.cell(1,6).value
    if (load_ws.cell(cnt,6).value != None):
        schoolname = load_ws.cell(cnt,6).value
        
    if (load_ws.cell(cnt,7).value == None): #비밀번호 있으면 그대로 쓰고 없으면 1행 
        password = load_ws.cell(1,7).value
    if (load_ws.cell(cnt,7).value != None): 
        password = load_ws.cell(cnt,7).value
        
    Cycle(load_ws.cell(cnt,1).value,load_ws.cell(cnt,2).value,city,level,schoolname,password) #이름, 생일, 시/도, 학교급, 학교이름, 비밀번호 값을 Cycle함수에 전달

for cnt in range(1,lastnumber) : #전부 시행하고 엑셀 3번열 삭제
    load_ws.cell(row=cnt,column=3).value = ""

print('EXCEL CLEARED')


load_wb.save(r'studentlist.xlsx')
