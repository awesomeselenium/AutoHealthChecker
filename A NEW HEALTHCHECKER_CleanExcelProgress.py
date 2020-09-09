from openpyxl import load_workbook
load_wb = load_workbook(r'studentlist.xlsx',data_only=True)
load_ws = load_wb['Sheet1']
lastnumber=1
while (load_ws.cell(lastnumber,1).value != None):
    lastnumber=lastnumber+1
for cnt in range(1,lastnumber) :
    load_ws.cell(row=cnt,column=3).value = ""
load_wb.save(r'studentlist.xlsx')

