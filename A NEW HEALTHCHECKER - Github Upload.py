#2020.09.07 Github Upload, 통합 자가진단 첫째날

from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import time

global driver

def START(): #webdriver 설정
    Options().add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36")
    global driver 
    driver = webdriver.Chrome('chromedriver')
    driver.get("https://hcs.eduro.go.kr/#/loginHome")
    time.sleep(1)
    
def AHC(name, birth,password):
    time.sleep(2)
    driver.implicitly_wait(10)
    driver.find_element_by_xpath('//*[@id="eventLayer"]/section[2]/button[2]').click() #앱설치 무시 웹으로 계속함
    driver.implicitly_wait(10)
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="btnConfirm2"]').click() #다음 페이지로 넘어감
    driver.implicitly_wait(10) 
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="WriteInfoForm"]/table/tbody/tr[2]/td/input').send_keys(name) #이름 입력 
    driver.find_element_by_xpath('//*[@id="WriteInfoForm"]/table/tbody/tr[3]/td/input').send_keys(birth) #생일 입력 
    driver.find_element_by_xpath('//*[@id="WriteInfoForm"]/table/tbody/tr[1]/td/button').click() #학교찾는 버튼 클릭 
    Select(driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[1]/table/tbody/tr[1]/td/select')).select_by_value("여기에 지역 코드 작성") #시/도 선택
    Select(driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[1]/table/tbody/tr[2]/td/select')).select_by_value("여기에 학교 급 작성") #학교급 선택
    driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[1]/table/tbody/tr[3]/td[1]/input').send_keys("여기에 학교 이름 작성") #학교 이름 작성
    driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[1]/table/tbody/tr[3]/td[2]/button').click() #학교 검색 버튼 클릭
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[1]/ul/li[1]/span').click() #학교 선택
    driver.find_element_by_xpath('//*[@id="softBoardListLayer"]/div[2]/div[2]/input').click()
    driver.find_element_by_xpath('//*[@id="btnConfirm"]').click() 
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="WriteInfoForm"]/table/tbody/tr/td/input').send_keys(password)#비밀번호 입력 
    driver.find_element_by_xpath('//*[@id="btnConfirm"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/section[2]/div[2]/ul/li/a/span[1]').click() #상위 사용자 선택 
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/div/div[2]/div[2]/dl[1]/dd/ul/li[1]/label').click() #1~5번 항목 확인
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/div/div[2]/div[2]/dl[2]/dd/ul/li[1]/label').click()
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/div/div[2]/div[2]/dl[3]/dd/ul/li[1]/label').click()
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/div/div[2]/div[2]/dl[4]/dd/ul/li[1]/label').click()
    driver.find_element_by_xpath('//*[@id="container"]/div[2]/div/div[2]/div[2]/dl[5]/dd/ul/li[1]/label').click()
    driver.find_element_by_xpath('//*[@id="btnConfirm"]').click() #제출 버튼 클릭
    time.sleep(1)
    print(load_ws.cell(cnt,1).value)
    print(cnt,"/",lastnumber-1,"completed")
    load_ws.cell(row=cnt,column=3).value = 1
    load_wb.save(r'studentlist.xlsx')

def RER(): #정보 클리어
    driver.refresh()
    time.sleep(1)
    driver.switch_to.alert.accept()
    time.sleep(1)
    #driver.switch_to.alert.accept()
    time.sleep(1)
    driver.find_element_by_xpath('/html/body/app-root/div/div[1]/div/button').click()    
    time.sleep(1)
    
    driver.switch_to.alert.accept()
    time.sleep(1)

def Cycle(name,birth,password): #이름,생일,비번 받아서 AHC와 RER을 수행
    AHC(name,birth,password)
    RER()
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
#프로그램 구동부
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
startnumber_manual=1
load_wb = load_workbook(r'studentlist.xlsx',data_only=True)
load_ws = load_wb['Sheet1']
lastnumber = 1
autostart = 1

START()

while (load_ws.cell(lastnumber,1).value != None): #엑셀 파일의 끝 확인
    lastnumber=lastnumber+1
while (load_ws.cell(autostart,3).value != None): #3열 이미 한 사람을 확인
    autostart=autostart+1
print("StartPoint:", autostart)
startnumber_manual= autostart
print("엑셀 파일의 학생은 자신의 건강상태와 주위 사람들의 건강 상태를 점검하여서 문제가 없었음을 인정하였습니다. ")

for cnt in range(startnumber_manual,lastnumber):
    if (load_ws.cell(cnt,4).value == None): #4번 열에 값 없으면 기본 비밀번호
        password ='2020' 
    if (load_ws.cell(cnt,4).value != None): #4번 열에 값 있으면 그 값 그대로 사용
        password = load_ws.cell(cnt,4).value
    Cycle(load_ws.cell(cnt,1).value,load_ws.cell(cnt,2).value,password) #이름, 생일, 비밀번호 값 함수에 전달

for cnt in range(1,lastnumber) : #3번 열을 확인용으로 사용하니 모든 확인이 끝난 후에 확인 열 클리어
    load_ws.cell(row=cnt,column=3).value = ""

load_wb.save(r'studentlist.xlsx')
